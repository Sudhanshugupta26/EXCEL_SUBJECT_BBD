import openpyxl
from openpyxl.styles import PatternFill
from supabase import create_client, Client

# Initialize Supabase Client
SUPABASE_URL = "https://tufsfubcdnwwafefhqko.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InR1ZnNmdWJjZG53d2FmZWZocWtvIiwicm9sZSI6ImFub24iLCJpYXQiOjE3MjQzNDU0OTcsImV4cCI6MjAzOTkyMTQ5N30.NA5pZBQpSZZrK3pECk3gPxphPU_RRbURK4Ct7DDH18Y"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Fetch data from Supabase
def fetch_papers():
    response = supabase.table('papers').select('subject_code, year, type').execute()
    return response.data

# Define the subjects and session types
subjects = [
    'COMMUNICATIVE ENGLISH','ENGINEERING CHEMISTRY','ENGINEERING PHYSICS','TECHNICAL COMMUNICATION','ENVIRONMENT & ECOLOGICAL SUSTAINABILITY','BASICS OF PYTHON PROGRAMMING','PROGRAMMING CONCEPTS WITH PYTHON','DATA VISUALIZATION WITH PYTHON', 'CHEMISTRY', 'AI IN MECHANICAL ENGINEERING SYSTEM', 'DIFFERENTIAL EQUATION AND FOURIER ANALYSIS', 'PHYSICS II', 'ENVIRONMENTAL STUDIES', 'COMPLEX ANALYSIS AND INTEGRAL TRANSFORMS', 'STATISTICAL AND NUMERICAL TECHNIQUES', 'DISCRETE MATHEMATICS', 'COMPUTER CONCEPTS AND PROGRAMMING IN C', 'DATABASE MANAGEMENT SYSTEMS', 'OPERATING SYSTEM', 'DATA STRUCTURE USING C', 'BASIC ELECTRONICS ENGINEERING', 'BASIC ELECTRICAL ENGINEERING', 'INDUSTRIAL PSYCHOLOGY', 'INDUSTRIAL SOCIOLOGY', 'ENGINEERING MECHANICS', 'INTERNET OF THINGS APPLICATION DEVELOPMENT', 'JAVA FUNDAMENTAL', 'BLOCKCHAIN ESSENTIALS', 'BUSINESS LAW', 'SALES MANAGEMENT', 'PRODUCTION AND OPERATIONS MANAGEMENT', 'MARKETING MANAGEMENT', 'HUMAN RESOURCE MANAGEMENT AND DEVELOPMENT', 'OPERATIONS RESEARCH', 'MATRICES AND CALCULAS', 'PHYSICS I', 'BASIC MECHANICAL ENGINEERING', 'PROFESSIONAL PRACTICE II', 'DEVELOPMENT LEGISLATIONS', 'ARCHITECTURAL DESIGN II', 'BUILDING CONSTRUCTION AND MATERIALS II', 'ARCHITECTURAL STRUCTURE II', 'ARCHITECTURAL DRAWING II', 'VISUAL ARTS II', 'HISTORY OF ARCHITECTURE AND CULTURE I', 'ARCHITECTURAL DESIGN IV', 'BUILDING CONSTRUCTION & MATERIALS IV', 'ARCHITECTURAL STRUCTURE IV', 'RESEARCH METHODOLOGY', 'ESSENTIALS OF BUSINESS COMMUNICATION', 'COMPUTER APPLICATION IN MANAGEMENT', 'BUSINESS STATISTICS', 'ORGANIZATIONAL BEHAVIOUR', 'DATABASE MANAGEMENT SYSTEM IMBA', 'CONSTRUCTION AND MATERIALS VIII', 'ARCHITECTURAL DESIGN VIII', 'BUILDING SERVICES III', 'SPECIFICATIONS AND ESTIMATION', 'CLIMATOLOGY', 'HISTORY OF ARCHITECTURE AND CULTURE III (ISLAMIC ARCH)', 'BUILDING SERVICES II', 'ARCHITECTURAL DESIGN VI', 'THEORY OF DESIGN', 'ARCHITECTURAL STRUCTURES VI', 'CONSTRUCTION AND MATERIALS VI', 'BUILDING SERVICES I', 'HISTORY OF ARCHITECTURE II', 'ARCHITECTURAL STRUCTURES IV', 'ARCHITECTURAL STRUCTURES II', 'BUILDING CONSTRUCTION & MATERIALS VI', 'SPECIFICATION COST ESTIMATION & BUDGETING', 'WORKING DRAWING', 'BUILDING SERVICES IV'
    # Add more subjects as needed
]
sessions = ['Sessional 1', 'Sessional 2', 'Semester', 'SCOP']

# Create Excel File
def create_excel_with_papers(data):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Previous Papers"

    # Add headers to the Excel sheet
    headers = ['Subject Name', 'Year', 'Sessional 1', 'Sessional 2', 'Semester', 'SCOP']
    ws.append(headers)

    # Define styles for green tick and red cross
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    tick = '✔'
    cross = '✘'

    # Insert subject, year, and availability of papers
    for subject in subjects:
        for year in range(2015, 2025):  # For each year between 2015 and 2024
            row = [subject, str(year)]
            for session in sessions:
                # Skip the first row in the database (data[1:] skips the first element)
                paper_found = any(
                    paper['subject_code'] == subject and paper['year'] == str(year) and paper['type'] == session
                    for paper in data[1:]  # Skip the first row
                )
                # Add green tick if found, red cross otherwise
                if paper_found:
                    row.append(tick)
                else:
                    row.append(cross)

            # Append the row to the worksheet
            ws.append(row)

            # Apply styling to the last four cells (Sessional 1, Sessional 2, Semester, SCOP)
            for col_idx in range(3, 7):  # Columns C to F correspond to session data including SCOP
                cell = ws.cell(row=len(ws['A']), column=col_idx)
                if cell.value == tick:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

    # Adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:  # Necessary to avoid errors if cell.value is None
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the Excel file
    wb.save("subject_papers_availability.xlsx")
    print("Excel file created successfully!")


if __name__ == "__main__":
    # Fetch data from Supabase
    papers_data = fetch_papers()

    # Check if data is fetched properly
    if not papers_data:
        raise ValueError("No data found from the database.")

    # Create Excel with fetched papers
    create_excel_with_papers(papers_data)
